"""
국외기업 추천 시스템 v2 - 협업필터링 SVD
추천기준: 업종코드유사 + 상품명유사(TF-IDF) + HS코드유사 + 등급(A최고) + 최신평가일 + 동일국가우선
"""
import pandas as pd
import numpy as np
from sklearn.decomposition import TruncatedSVD
from sklearn.feature_extraction.text import TfidfVectorizer
from sklearn.metrics.pairwise import cosine_similarity
from sklearn.preprocessing import normalize
from scipy.sparse import hstack, lil_matrix
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import warnings
warnings.filterwarnings('ignore')

# ════════════════════════════════════════════════════════
# 설정
# ════════════════════════════════════════════════════════
TOP_N = 10
SVD_COMPONENTS = 80

# 유사도 구성 가중치 (특성행렬 내 블록 가중치)
FEAT_W = {'업종코드': 2.5, 'HS코드': 2.5, '상품명': 2.0}

# 최종 점수 가중치
SCORE_W = {'svd': 0.40, 'grade': 0.28, 'recency': 0.12, 'country': 0.20}

GRADE_SCORE = {
    'A': 13, 'B': 12, 'C1': 11, 'C2': 10,
    'D1': 9,  'D2': 8,  'E1': 7,  'E2': 6,
    'F1': 5,  'F2': 4,  'G1': 3,  'G2': 2, 'R': 1
}
SCORE_TO_GRADE = {v: k for k, v in GRADE_SCORE.items()}

OUTPUT_PATH = '/home/claude/국외기업_추천결과_v2.xlsx'

# ════════════════════════════════════════════════════════
# 1. 데이터 로드 & 전처리
# ════════════════════════════════════════════════════════
print("▶ [1/6] 데이터 로딩 및 전처리...")
input_df = pd.read_excel('/mnt/user-data/uploads/입력데이터.xlsx')
ref_df   = pd.read_excel('/mnt/user-data/uploads/참고데이터.xlsx')

def preprocess(df):
    df = df.copy()
    df['수입자코드']    = df['수입자코드'].astype(str).str.strip()
    df['국가코드']      = df['국가코드'].fillna(0).astype(float).astype(int).astype(str)
    df['국가명']        = df['국가명'].fillna('').astype(str)

    # 업종코드 계층 분해 (앞 3자리 / 앞 2자리)
    biz = df['업종코드'].fillna(0).astype(float).astype(int).astype(str).str.zfill(5)
    df['업종_3'] = biz.str[:3]
    df['업종_2'] = biz.str[:2]

    # HS코드 계층 분해 (앞 4자리 / 앞 2자리)
    hs = pd.to_numeric(df['HS코드'], errors='coerce').fillna(0).astype(int).astype(str).str.zfill(10)
    df['HS_4'] = hs.str[:4]
    df['HS_2'] = hs.str[:2]

    # 상품명 + 업종명 텍스트 결합 (TF-IDF용)
    prod = df['상품명'].fillna('').astype(str)
    biz_nm = df['업종명'].fillna('').astype(str) if '업종명' in df.columns else ''
    df['text'] = (prod + ' ' + biz_nm).str.strip().str.lower()

    # 평가 관련
    df['평가일자_num'] = pd.to_numeric(df['평가일자'], errors='coerce').fillna(0).astype(int)
    df['등급_점수']    = df['평가등급'].map(GRADE_SCORE).fillna(0)
    return df

input_df = preprocess(input_df)
ref_df   = preprocess(ref_df)

# 입력 수입자코드 제외
input_codes = set(input_df['수입자코드'].unique())
ref_df = ref_df[~ref_df['수입자코드'].isin(input_codes)].copy().reset_index(drop=True)
print(f"   입력: {input_df['수입자코드'].nunique()}개 기업 | "
      f"참고: {ref_df['수입자코드'].nunique():,}개 기업 (입력코드 제외 후)")

# ════════════════════════════════════════════════════════
# 2. 수입자 단위 집계
# ════════════════════════════════════════════════════════
print("▶ [2/6] 수입자 단위 집계...")

def aggregate(df):
    agg = df.groupby('수입자코드').agg(
        수입자명        = ('수입자명',     'first'),
        국가코드        = ('국가코드',     'first'),
        국가명          = ('국가명',       'first'),
        업종_3_list     = ('업종_3',       lambda x: list(set(x))),
        업종_2_list     = ('업종_2',       lambda x: list(set(x))),
        HS_4_list       = ('HS_4',         lambda x: list(set(x))),
        HS_2_list       = ('HS_2',         lambda x: list(set(x))),
        text            = ('text',         lambda x: ' '.join(set(x))),
        최고_등급점수   = ('등급_점수',    'max'),
        최근_평가일자   = ('평가일자_num', 'max'),
    ).reset_index()

    agg['최고_평가등급'] = agg['최고_등급점수'].map(SCORE_TO_GRADE).fillna('-')

    # 최고 등급 행 기준 대표 정보 추출
    best_rows = (
        df.sort_values('등급_점수', ascending=False)
          .groupby('수입자코드')
          .first()
          .reset_index()
        [['수입자코드', '업종코드', '업종명', 'HS코드', 'HS코드명', '대표자명']]
    )
    best_rows.columns = ['수입자코드','대표_업종코드','대표_업종명','대표_HS코드','대표_HS코드명','대표자명']
    agg = agg.merge(best_rows, on='수입자코드', how='left')
    return agg

input_agg = aggregate(input_df)
ref_agg   = aggregate(ref_df)
n_inp = len(input_agg)
n_ref = len(ref_agg)
print(f"   집계 완료 — 입력: {n_inp}개 | 참고: {n_ref:,}개")

# ════════════════════════════════════════════════════════
# 3. 특성 행렬 구성 (업종코드 + HS코드 + 상품명 TF-IDF)
# ════════════════════════════════════════════════════════
print("▶ [3/6] 특성 행렬 구성 (업종코드 + HS코드 + TF-IDF)...")
all_agg = pd.concat([input_agg, ref_agg], ignore_index=True)
N = len(all_agg)

# ── 블록 1: 업종코드 멀티-핫 (앞3자리 × 2.5 + 앞2자리 × 1.5) ──
def build_multihot(agg, col_list, weight=1.0):
    vocab = sorted({v for lst in agg[col_list] for v in lst})
    idx   = {v: i for i, v in enumerate(vocab)}
    mat   = lil_matrix((N, len(vocab)), dtype=np.float32)
    for i, lst in enumerate(agg[col_list]):
        for v in lst:
            if v in idx:
                mat[i, idx[v]] = weight
    return mat.tocsr(), vocab

mat_업종3, _ = build_multihot(all_agg, '업종_3_list', FEAT_W['업종코드'])
mat_업종2, _ = build_multihot(all_agg, '업종_2_list', FEAT_W['업종코드'] * 0.6)
mat_HS4,   _ = build_multihot(all_agg, 'HS_4_list',   FEAT_W['HS코드'])
mat_HS2,   _ = build_multihot(all_agg, 'HS_2_list',   FEAT_W['HS코드'] * 0.5)

# ── 블록 2: 상품명 TF-IDF ──
texts = all_agg['text'].fillna('').tolist()
tfidf = TfidfVectorizer(
    max_features=3000,
    ngram_range=(1, 2),
    min_df=2,
    sublinear_tf=True,
    strip_accents='unicode',
    token_pattern=r'(?u)\b[a-zA-Z][a-zA-Z0-9]{1,}\b'
)
mat_tfidf = tfidf.fit_transform(texts) * FEAT_W['상품명']

# ── 결합 ──
feat_matrix = hstack([mat_업종3, mat_업종2, mat_HS4, mat_HS2, mat_tfidf], format='csr')
print(f"   특성 행렬: {feat_matrix.shape} | nnz: {feat_matrix.nnz:,}")

# ════════════════════════════════════════════════════════
# 4. TruncatedSVD → 잠재 공간 코사인 유사도
# ════════════════════════════════════════════════════════
print("▶ [4/6] SVD 분해 및 유사도 계산...")
n_comp = min(SVD_COMPONENTS, feat_matrix.shape[1] - 1, N - 1)
svd = TruncatedSVD(n_components=n_comp, n_iter=15, random_state=42)
latent = svd.fit_transform(feat_matrix)
latent = normalize(latent, norm='l2')  # 코사인 유사도 위한 L2 정규화
print(f"   SVD 컴포넌트: {n_comp} | 설명 분산: {svd.explained_variance_ratio_.sum():.2%}")

inp_latent = latent[:n_inp]
ref_latent = latent[n_inp:]

# 배치 코사인 유사도 (내적 = L2-정규화 후 점곱)
sim_matrix = inp_latent @ ref_latent.T  # (n_inp, n_ref)

# ════════════════════════════════════════════════════════
# 5. 보조 점수 정규화
# ════════════════════════════════════════════════════════
ref_dates  = ref_agg['최근_평가일자'].values.astype(float)
ref_grades = ref_agg['최고_등급점수'].values.astype(float)

date_min, date_max = ref_dates.min(), ref_dates.max()
recency_score = (ref_dates - date_min) / max(date_max - date_min, 1)
grade_score   = ref_grades / 13.0

# ════════════════════════════════════════════════════════
# 6. 추천 생성
# ════════════════════════════════════════════════════════
print(f"▶ [5/6] TOP-{TOP_N} 추천 생성...")
results = []

for i, inp_row in input_agg.iterrows():
    sims    = sim_matrix[i]
    country_bonus = (ref_agg['국가코드'].values == inp_row['국가코드']).astype(float)

    final = (
        SCORE_W['svd']     * sims +
        SCORE_W['grade']   * grade_score +
        SCORE_W['recency'] * recency_score +
        SCORE_W['country'] * country_bonus
    )

    top_idx = np.argsort(final)[::-1][:TOP_N]

    for rank, ri in enumerate(top_idx, 1):
        rr = ref_agg.iloc[ri]
        d  = str(int(rr['최근_평가일자']))
        date_fmt = f"{d[:4]}-{d[4:6]}-{d[6:8]}" if len(d) == 8 else d

        results.append({
            '입력_수입자코드':   inp_row['수입자코드'],
            '입력_수입자명':     inp_row['수입자명'],
            '입력_국가명':       inp_row['국가명'],
            '추천순위':          rank,
            '추천_수입자코드':   rr['수입자코드'],
            '추천_수입자명':     rr['수입자명'],
            '추천_국가명':       rr['국가명'],
            '추천_대표자명':     rr['대표자명'],
            '추천_업종코드':     rr['대표_업종코드'],
            '추천_업종명':       rr['대표_업종명'],
            '추천_HS코드':       rr['대표_HS코드'],
            '추천_HS코드명':     rr['대표_HS코드명'],
            '추천_상품명':       rr['text'][:120],
            '추천_최고평가등급': rr['최고_평가등급'],
            '추천_최근평가일자': date_fmt,
            '국가동일여부':      '✔ 동일' if country_bonus[ri] else '다름',
            'SVD_유사도':        round(float(sims[ri]), 4),
            '등급점수(정규화)':  round(float(grade_score[ri]), 4),
            '최신성점수':        round(float(recency_score[ri]), 4),
            '국가보너스':        round(float(country_bonus[ri]), 4),
            '최종_추천점수':     round(float(final[ri]), 4),
        })

result_df = pd.DataFrame(results)
print(f"   추천 레코드: {len(result_df):,}행")

# ════════════════════════════════════════════════════════
# 7. 엑셀 출력
# ════════════════════════════════════════════════════════
print(f"▶ [6/6] 엑셀 파일 생성: {OUTPUT_PATH}")

# ── 스타일 헬퍼 ──
NAVY      = "1F3864"
WHITE     = "FFFFFF"
GOLD_BG   = "FFF2CC"
SILVER_BG = "F2F2F2"
GREEN_BG  = "E2EFDA"
BLUE_BG   = "D6E4F7"
STRIPE    = "F7FBFF"

GRADE_CFG = {
    'A':  ('C6EFCE', '375623', True),
    'B':  ('DDEBF7', '1F3864', True),
    'C1': ('FFEB9C', '7D6608', False),
    'C2': ('FFF2CC', '7D6608', False),
    'D1': ('FCE4D6', '843C0C', False),
    'D2': ('FCE4D6', '843C0C', False),
    'E1': ('F4CCCC', '9C0006', False),
    'E2': ('F4CCCC', '9C0006', False),
    'F1': ('EA9999', '9C0006', True),
    'F2': ('EA9999', '9C0006', True),
    'G1': ('CC0000', 'FFFFFF', True),
    'G2': ('990000', 'FFFFFF', True),
}

def thin_border(color="B8CCE4"):
    s = Side(style='thin', color=color)
    return Border(left=s, right=s, top=s, bottom=s)

BD = thin_border()

def hdr(ws, row, col, val, width=None, wrap=True):
    c = ws.cell(row=row, column=col, value=val)
    c.font      = Font(name='맑은 고딕', bold=True, color=WHITE, size=10)
    c.fill      = PatternFill('solid', fgColor=NAVY)
    c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=wrap)
    c.border    = BD
    if width:
        ws.column_dimensions[get_column_letter(col)].width = width
    return c

def cell(ws, row, col, val, bg=None, bold=False, align='center', color='000000', size=9):
    c = ws.cell(row=row, column=col, value=val)
    c.font      = Font(name='맑은 고딕', bold=bold, size=size, color=color)
    c.alignment = Alignment(horizontal=align, vertical='center', wrap_text=True)
    c.border    = BD
    if bg:
        c.fill = PatternFill('solid', fgColor=bg)
    return c

def title_row(ws, text, cols, row=1, h=32, size=13):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cols)
    c = ws.cell(row=row, column=1, value=text)
    c.font      = Font(name='맑은 고딕', bold=True, size=size, color=WHITE)
    c.fill      = PatternFill('solid', fgColor=NAVY)
    c.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[row].height = h

wb = openpyxl.Workbook()

# ══════════════════════════════════════════════════════
# Sheet 1 : 추천결과_전체
# ══════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = '추천결과_전체'

COLS1 = [
    ('입력_수입자코드',   13), ('입력_수입자명',     26), ('입력_국가명',   10),
    ('추천\n순위',         6), ('추천_수입자코드',   13), ('추천_수입자명', 30),
    ('추천_국가명',       10), ('추천_대표자명',     16), ('추천_업종코드',  8),
    ('추천_업종명',       22), ('추천_HS코드',       12), ('추천_HS코드명', 18),
    ('추천_상품명(요약)', 32), ('최고\n평가등급',     8), ('최근\n평가일자',11),
    ('국가\n동일여부',     8), ('SVD\n유사도',        8), ('등급\n점수',     8),
    ('최신성\n점수',       8), ('국가\n보너스',       8), ('최종\n추천점수',10),
]
title_row(ws1, '국외기업 SVD 협업필터링 추천 결과 (업종코드 + 상품명 + HS코드 유사도 기반)', len(COLS1), row=1)
ws1.row_dimensions[2].height = 36
ws1.freeze_panes = 'A3'

for ci, (name, w) in enumerate(COLS1, 1):
    hdr(ws1, 2, ci, name, width=w)

DATA_COLS = [
    '입력_수입자코드','입력_수입자명','입력_국가명',
    '추천순위','추천_수입자코드','추천_수입자명',
    '추천_국가명','추천_대표자명','추천_업종코드',
    '추천_업종명','추천_HS코드','추천_HS코드명',
    '추천_상품명','추천_최고평가등급','추천_최근평가일자',
    '국가동일여부','SVD_유사도','등급점수(정규화)',
    '최신성점수','국가보너스','최종_추천점수',
]
LEFT_COLS = {2, 6, 10, 12, 13}  # 좌정렬

prev_input = None
stripe = False
for ri, row in result_df.iterrows():
    er = ri + 3
    ws1.row_dimensions[er].height = 18

    # 입력기업 구분선 (다른 입력기업 시작 시 배경 교체)
    if row['입력_수입자코드'] != prev_input:
        stripe = not stripe
        prev_input = row['입력_수입자코드']

    rank = row['추천순위']
    same = row['국가동일여부'] == '✔ 동일'

    if rank == 1:
        row_bg = GOLD_BG
    elif rank == 2:
        row_bg = SILVER_BG
    elif same:
        row_bg = GREEN_BG
    elif stripe:
        row_bg = STRIPE
    else:
        row_bg = None

    for ci, col_name in enumerate(DATA_COLS, 1):
        val   = row[col_name]
        align = 'left' if ci in LEFT_COLS else 'center'
        c     = cell(ws1, er, ci, val, bg=row_bg, align=align)

    # 등급 셀 색상 덮어쓰기
    grade = str(row['추천_최고평가등급'])
    if grade in GRADE_CFG:
        bg_g, fg_g, bold_g = GRADE_CFG[grade]
        gc = ws1.cell(row=er, column=14)
        gc.fill = PatternFill('solid', fgColor=bg_g)
        gc.font = Font(name='맑은 고딕', bold=bold_g, size=9, color=fg_g)

# ══════════════════════════════════════════════════════
# Sheet 2 : TOP5_요약
# ══════════════════════════════════════════════════════
ws2 = wb.create_sheet('TOP5_요약')
COLS2 = [
    ('입력_수입자코드',   13), ('입력_수입자명',     26), ('입력_국가명',  10),
    ('추천\n순위',         6), ('추천_수입자코드',   13), ('추천_수입자명',30),
    ('추천_국가명',       10), ('추천_업종명',       22), ('추천_상품명(요약)',30),
    ('최고\n평가등급',     8), ('최근\n평가일자',    11), ('국가\n동일여부', 8),
    ('최종\n추천점수',    10),
]
title_row(ws2, 'TOP-5 추천 요약 (입력기업별)', len(COLS2), row=1)
ws2.row_dimensions[2].height = 36
ws2.freeze_panes = 'A3'
for ci, (name, w) in enumerate(COLS2, 1):
    hdr(ws2, 2, ci, name, width=w)

top5 = result_df[result_df['추천순위'] <= 5].reset_index(drop=True)
DATA2 = [
    '입력_수입자코드','입력_수입자명','입력_국가명',
    '추천순위','추천_수입자코드','추천_수입자명',
    '추천_국가명','추천_업종명','추천_상품명',
    '추천_최고평가등급','추천_최근평가일자','국가동일여부',
    '최종_추천점수',
]
LEFT2 = {2, 6, 8, 9}
prev_input = None
stripe = False
for ri, row in top5.iterrows():
    er = ri + 3
    ws2.row_dimensions[er].height = 18
    if row['입력_수입자코드'] != prev_input:
        stripe = not stripe
        prev_input = row['입력_수입자코드']
    rank = row['추천순위']
    same = row['국가동일여부'] == '✔ 동일'
    if rank == 1:   row_bg = GOLD_BG
    elif rank == 2: row_bg = SILVER_BG
    elif same:      row_bg = GREEN_BG
    elif stripe:    row_bg = STRIPE
    else:           row_bg = None
    for ci, col_name in enumerate(DATA2, 1):
        align = 'left' if ci in LEFT2 else 'center'
        cell(ws2, er, ci, row[col_name], bg=row_bg, align=align)
    grade = str(row['추천_최고평가등급'])
    if grade in GRADE_CFG:
        bg_g, fg_g, bold_g = GRADE_CFG[grade]
        gc = ws2.cell(row=er, column=10)
        gc.fill = PatternFill('solid', fgColor=bg_g)
        gc.font = Font(name='맑은 고딕', bold=bold_g, size=9, color=fg_g)

# ══════════════════════════════════════════════════════
# Sheet 3 : 입력기업별 1위 현황
# ══════════════════════════════════════════════════════
ws3 = wb.create_sheet('1위_추천현황')
top1 = result_df[result_df['추천순위'] == 1].reset_index(drop=True)
COLS3 = [
    ('입력_수입자코드', 13), ('입력_수입자명', 26), ('입력_국가명', 10),
    ('추천_수입자코드', 13), ('추천_수입자명', 30), ('추천_국가명', 10),
    ('추천_업종코드',    8), ('추천_업종명',   22), ('추천_상품명(요약)', 32),
    ('추천_HS코드',     12), ('최고\n평가등급', 8), ('최근\n평가일자',  11),
    ('국가\n동일여부',   8), ('최종\n추천점수',10),
]
title_row(ws3, '입력기업별 1위 추천 현황', len(COLS3), row=1)
ws3.row_dimensions[2].height = 36
ws3.freeze_panes = 'A3'
for ci, (name, w) in enumerate(COLS3, 1):
    hdr(ws3, 2, ci, name, width=w)
DATA3 = [
    '입력_수입자코드','입력_수입자명','입력_국가명',
    '추천_수입자코드','추천_수입자명','추천_국가명',
    '추천_업종코드','추천_업종명','추천_상품명',
    '추천_HS코드','추천_최고평가등급','추천_최근평가일자',
    '국가동일여부','최종_추천점수',
]
LEFT3 = {2, 5, 8, 9}
for ri, row in top1.iterrows():
    er = ri + 3
    ws3.row_dimensions[er].height = 18
    same = row['국가동일여부'] == '✔ 동일'
    row_bg = GREEN_BG if same else (STRIPE if ri % 2 == 0 else None)
    for ci, col_name in enumerate(DATA3, 1):
        align = 'left' if ci in LEFT3 else 'center'
        cell(ws3, er, ci, row[col_name], bg=row_bg, align=align)
    grade = str(row['추천_최고평가등급'])
    if grade in GRADE_CFG:
        bg_g, fg_g, bold_g = GRADE_CFG[grade]
        gc = ws3.cell(row=er, column=11)
        gc.fill = PatternFill('solid', fgColor=bg_g)
        gc.font = Font(name='맑은 고딕', bold=bold_g, size=9, color=fg_g)

# ══════════════════════════════════════════════════════
# Sheet 4 : 알고리즘 설명
# ══════════════════════════════════════════════════════
ws4 = wb.create_sheet('알고리즘_설명')
title_row(ws4, '국외기업 SVD 협업필터링 추천 알고리즘 설명', 3, row=1)

INFO = [
    ['구분', '항목', '상세 내용'],
    ['특성 구성', '업종코드 유사도',
     f'업종코드 앞3자리(가중치 {FEAT_W["업종코드"]}) + 앞2자리(가중치 {FEAT_W["업종코드"]*0.6}) → 멀티-핫 인코딩'],
    ['특성 구성', 'HS코드 유사도',
     f'HS코드 앞4자리(가중치 {FEAT_W["HS코드"]}) + 앞2자리(가중치 {FEAT_W["HS코드"]*0.5}) → 멀티-핫 인코딩'],
    ['특성 구성', '상품명 유사도',
     f'상품명+업종명 텍스트 → TF-IDF (최대 3,000 토큰, 1~2-gram, 가중치 {FEAT_W["상품명"]}) → 통합 행렬'],
    ['SVD 분해', 'TruncatedSVD',
     f'3가지 특성 행렬 수평 결합 후 잠재 공간 분해 (컴포넌트: {n_comp}, 설명분산: {svd.explained_variance_ratio_.sum():.2%})'],
    ['최종 점수', 'SVD 유사도',
     f'잠재 공간 코사인 유사도 (가중치 {SCORE_W["svd"]*100:.0f}%)'],
    ['최종 점수', '등급 점수',
     f'A=13점 ~ R=1점 정규화 (가중치 {SCORE_W["grade"]*100:.0f}%) — A등급이 최고 우선순위'],
    ['최종 점수', '최신성 점수',
     f'최근 평가일자를 0~1로 정규화 (가중치 {SCORE_W["recency"]*100:.0f}%)'],
    ['최종 점수', '국가 동일 보너스',
     f'입력 수입자와 동일 국가 = 1.0, 다른 국가 = 0.0 (가중치 {SCORE_W["country"]*100:.0f}%)'],
    ['등급 체계', '등급 순위',
     'A > B > C1 > C2 > D1 > D2 > E1 > E2 > F1 > F2 > G1 > G2 > R'],
    ['출력 현황', '처리 결과',
     f'입력 기업 {n_inp}개 × TOP-{TOP_N} = {len(result_df):,}개 추천 레코드'],
]

ws4.column_dimensions['A'].width = 14
ws4.column_dimensions['B'].width = 22
ws4.column_dimensions['C'].width = 80
for ri, row_data in enumerate(INFO, 2):
    ws4.row_dimensions[ri].height = 24
    for ci, val in enumerate(row_data, 1):
        is_hdr = ri == 2
        c = ws4.cell(row=ri, column=ci, value=val)
        c.font      = Font(name='맑은 고딕', bold=is_hdr or ci==1, size=10,
                           color=WHITE if is_hdr else '000000')
        c.fill      = PatternFill('solid', fgColor=NAVY if is_hdr
                                  else ('D6E4F7' if ri % 2 == 0 else 'EBF3FB'))
        c.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        c.border    = BD

# 범례
LG_ROW = len(INFO) + 3
cell(ws4, LG_ROW,   1, '등급별 색상 범례', bold=True, bg=NAVY, color=WHITE)
cell(ws4, LG_ROW,   2, '',                  bg=NAVY)
cell(ws4, LG_ROW,   3, '',                  bg=NAVY)
ws4.row_dimensions[LG_ROW].height = 20
for j, (grade, (bg_g, fg_g, bold_g)) in enumerate(list(GRADE_CFG.items()), 0):
    r = LG_ROW + 1 + (j // 4)
    c_col = (j % 4) + 1
    gc = ws4.cell(row=r, column=c_col, value=f"  등급 {grade}")
    gc.fill = PatternFill('solid', fgColor=bg_g)
    gc.font = Font(name='맑은 고딕', bold=bold_g, size=10, color=fg_g)
    gc.alignment = Alignment(horizontal='center', vertical='center')
    gc.border = BD
    ws4.row_dimensions[r].height = 20

wb.save(OUTPUT_PATH)
print(f"✅ 완료: {OUTPUT_PATH}")
print(f"   시트 1 추천결과_전체 : {len(result_df):,}행")
print(f"   시트 2 TOP5_요약     : {len(top5):,}행")
print(f"   시트 3 1위_추천현황  : {len(top1)}행")
print(f"   시트 4 알고리즘_설명 : 파라미터 및 등급 범례")
